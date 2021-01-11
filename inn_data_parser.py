import requests
import re
from datetime import datetime, timedelta
import json
from fake_useragent import UserAgent
import openpyxl
from lxml.html import fromstring
from itertools import cycle
import traceback
import time
from selenium import webdriver

#Вводные данные
#Дравер
DRIVER = 'chromedriver'
driver = webdriver.Chrome('chromedriver.exe')
#В корне программы должны лежать 2 файла: input_inn.xlsx (в первом столбце ИНН), output_inn.xlsx - файл для записи
input_inn = openpyxl.load_workbook(filename = 'input_inn.xlsx')
output_inn = openpyxl.load_workbook(filename = 'output_inn.xlsx')

sheet_input = input_inn.worksheets[0]
sheet = output_inn.worksheets[0]
row_count = sheet_input.max_row
column_count = sheet_input.max_column

#Прокси (не потребовалось)
def get_proxies():
    url = 'https://free-proxy-list.net/'
    response = requests.get(url)
    parser = fromstring(response.text)
    proxies = set()
    for i in parser.xpath('//tbody/tr')[:500]:
        if i.xpath('.//td[7][contains(text(),"yes")]'):
            proxy = ":".join([i.xpath('.//td[1]/text()')[0], i.xpath('.//td[2]/text()')[0]])
            proxies.add(proxy)
    return proxies


def get_load_json(inn,i,row_count):
    time.sleep(3) #Пауза обращения
    ua = UserAgent()
    headers = {'User-Agent': ua.random}
    url = 'https://www.rusprofile.ru/ajax.php?&query={0}&action=search'.format(inn)
    profile = requests.get(url).json()
    try:
        profile = profile['ul']
        write_excel_file(profile,i)
        print('Данные по: ',inn,' найдены.  Выполнено: ',i,'/',row_count)
        sheet_input.cell(row=i, column=1).value = 0
        input_inn.save('input_inn.xlsx')
    except BaseException:
        print('Данные по: ',inn,' НЕ НАЙДЕНЫ (Сделано слишком много запросов. Нажми на капчу)')
        driver.get('https://www.rusprofile.ru/')
        vbr3 = input("Повторить ? y/n ")
        if vbr3 == 'y':
            get_load_json(inn,i,row_count)  
            
    
def write_excel_file(profile,i):
    #Создаем в Excel файле заголовки 
    sheet.cell(row=1, column=1).value = 'name'
    sheet.cell(row=1, column=2).value = 'raw_name'
    sheet.cell(row=1, column=3).value = 'many_ceo'
    sheet.cell(row=1, column=4).value = 'link'
    sheet.cell(row=1, column=5).value = 'ogrn'
    sheet.cell(row=1, column=6).value = 'raw_ogrn'
    sheet.cell(row=1, column=7).value = 'inn'
    sheet.cell(row=1, column=8).value = 'region'
    sheet.cell(row=1, column=9).value = 'address'
    sheet.cell(row=1, column=10).value = 'inactive'
    sheet.cell(row=1, column=11).value = 'status_extended'
    sheet.cell(row=1, column=12).value = 'ceo_name'
    sheet.cell(row=1, column=13).value = 'ceo_type'
    sheet.cell(row=1, column=14).value = 'snippet_string'
    sheet.cell(row=1, column=15).value = 'snippet_type'
    sheet.cell(row=1, column=16).value = 'status_code'
    sheet.cell(row=1, column=17).value = 'svprekrul_date'
    sheet.cell(row=1, column=18).value = 'main_okved_id'
    sheet.cell(row=1, column=19).value = 'okved_descr'
    sheet.cell(row=1, column=20).value = 'authorized_capital'
    sheet.cell(row=1, column=21).value = 'reg_date'
    sheet.cell(row=1, column=22).value = 'url'
    #Записываем данные построчно
    sheet.cell(row=i+1, column=1).value = profile[0]['name']
    sheet.cell(row=i+1, column=2).value = profile[0]['raw_name']
    sheet.cell(row=i+1, column=3).value = profile[0]['many_ceo']
    sheet.cell(row=i+1, column=4).value = 'https://www.rusprofile.ru' + profile[0]['link']
    sheet.cell(row=i+1, column=5).value = profile[0]['ogrn']
    sheet.cell(row=i+1, column=6).value = profile[0]['raw_ogrn']
    sheet.cell(row=i+1, column=7).value = profile[0]['inn'].replace('~','').replace('!','')
    sheet.cell(row=i+1, column=8).value = profile[0]['region']
    sheet.cell(row=i+1, column=9).value = profile[0]['address']
    sheet.cell(row=i+1, column=10).value = profile[0]['inactive']
    sheet.cell(row=i+1, column=11).value = profile[0]['status_extended']
    sheet.cell(row=i+1, column=12).value = profile[0]['ceo_name']
    sheet.cell(row=i+1, column=13).value = profile[0]['ceo_type']
    sheet.cell(row=i+1, column=14).value = profile[0]['snippet_string']
    sheet.cell(row=i+1, column=15).value = profile[0]['snippet_type']
    sheet.cell(row=i+1, column=16).value = profile[0]['status_code']
    sheet.cell(row=i+1, column=17).value = profile[0]['svprekrul_date']
    sheet.cell(row=i+1, column=18).value = profile[0]['main_okved_id']
    sheet.cell(row=i+1, column=19).value = profile[0]['okved_descr']
    sheet.cell(row=i+1, column=20).value = profile[0]['authorized_capital']
    sheet.cell(row=i+1, column=21).value = profile[0]['reg_date']
    sheet.cell(row=i+1, column=22).value = 'https://www.rusprofile.ru' + profile[0]['url']    
    output_inn.save('output_inn.xlsx') 

#Создание текстового файла 
def create_txt_file():
    text_file = open('output.txt','w',encoding='utf-8')
    text_file.write(str('name|raw_name|many_ceo|link|name|ogrn|raw_ogrn|inn|region|address|inactive|status_extended|ceo_name|ceo_type|snippet_string|snippet_type|status_code|svprekrul_date|main_okved_id|okved_descr|authorized_capital|reg_date|url'))
    text_file.write('\n')
    text_file.close()

#Запись в текстовый файл
def write_txt_file(company):
    text_file = open('output.txt','a',encoding='utf-8')
    text_file.write('{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}|{}\n'.format(str(company['name']),str(company['raw_name']),str(company['many_ceo']),str(company['link']),str(company['ogrn']),str(company['raw_ogrn']),\
                        str(company['inn']),str(company['region']),str(company['address']),str(company['inactive']),str(company['status_extended']),str(company['ceo_name']),\
                        str(company['ceo_type']),str(company['snippet_string']),str(company['snippet_type']),str(company['status_code']),str(company['svprekrul_date']),\
                        str(company['main_okved_id']),str(company['okved_descr']),str(company['authorized_capital']),str(company['reg_date']),str(company['url'])))
    text_file.close()    
    
vbr1 = input("Вы готовы искать? y/n ?)")
if vbr1 == 'y':
    #proxies = get_proxies()
    #proxy_pool = cycle(proxies)
    for i in range(1,row_count+1):
#        proxy = next(proxy_pool)
        inn = sheet_input.cell(row=i, column=1).value
        if len(str(inn)) == 9: #Если ИНН без нуля в [0] 
            inn = '0'+str(inn)
            int(inn)
            get_load_json(inn,i,row_count)
        if len(str(inn)) == 10:
            get_load_json(inn,i,row_count)

        
vbr2 = input("Сохранить и выйти? y/n ?)")
if vbr2 == 'y':  
    output_inn.save('output_inn.xlsx')
